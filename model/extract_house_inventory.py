#!/usr/bin/env python3
"""Build canonical house inventory artifacts from spreadsheets + scanned drawing decisions.

Usage:
  uv run --with openpyxl python model/extract_house_inventory.py

Outputs (model/data/inventory/):
  - house_inventory.json            # canonical machine-readable inventory
  - heating_demand_items.csv        # row-level Heating Demand table extract
  - room_dimension_summary.csv      # per-room geometry/envelope summary
  - scan_dimension_points.csv       # manually transcribed dimensions from scans
  - radiator_inventory_current.csv  # current radiator table extract
  - radiator_inventory_change.csv   # proposed/change radiator table extract
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
HEATING_XLSX = ROOT / "Heating needs for the house.xlsx"
LOFT_HEATING_XLSX = Path("/mnt/c/Users/jackc/OneDrive/Documents/House/Heating needs for the loft.xlsx")
OUT_DIR = ROOT / "model" / "data" / "inventory"
THERMAL_GEOMETRY_JSON = ROOT / "data" / "canonical" / "thermal_geometry.json"
SUNLIGHT_SCAN_DIR = Path(
    "/mnt/c/Users/jackc/AppData/Roaming/PFU/ScanSnap Home/ScanSnap Home"
)
SUNLIGHT_OVERVIEW = Path("/mnt/c/Users/jackc/OneDrive/IMG_0829.jpeg")
BCA_SECTION_PDF = Path("/mnt/c/Users/jackc/OneDrive/Documents/House/2528-08 PROPOSED EXTENSION SECTIONS.pdf")
BCA_EXISTING_PDF = Path("/mnt/c/Users/jackc/OneDrive/Documents/House/2528-02 EXISTING ELEVATIONS.pdf")


def _norm(v: Any) -> Any:
    if isinstance(v, str):
        return v.strip()
    return v


def _s(v: Any) -> str:
    return "" if v is None else str(v)


def canonical_room_name(name: str) -> str:
    return {
        "Carol Bedrrom": "Jack & Carol",
        "Front Room": "Front",
        "Consevatory": "Conservatory",
    }.get(name, name)


def find_headers(ws, first_col_text: str, second_col_prefix: str) -> list[int]:
    rows: list[int] = []
    for r in range(1, ws.max_row + 1):
        c1 = _norm(ws.cell(r, 1).value)
        c2 = _norm(ws.cell(r, 2).value)
        c2s = c2 if isinstance(c2, str) else ""
        if c1 == first_col_text and c2s.startswith(second_col_prefix):
            rows.append(r)
    return rows


def parse_heating_demand(ws) -> tuple[list[dict[str, Any]], dict[str, dict[str, float]]]:
    header_rows = find_headers(ws, "Room", "Item")
    if not header_rows:
        raise RuntimeError("Could not find Heating Demand header row")
    h = header_rows[0]

    cols = {
        "room": 1,
        "item": 2,
        "meters": 3,
        "ceiling_height": 4,
        "area_m2": 5,
        "u_value": 6,
        "temp_diff_c": 7,
        "watts": 8,
        "kwh_per_year": 9,
    }

    rows: list[dict[str, Any]] = []
    room_totals: dict[str, dict[str, float]] = {}

    r = h + 1
    while r <= ws.max_row:
        room = _norm(ws.cell(r, cols["room"]).value)
        if room in (None, "", "Total"):
            break
        rec = {k: _norm(ws.cell(r, c).value) for k, c in cols.items()}
        rec["source"] = {"file": HEATING_XLSX.name, "sheet": ws.title, "row": r}

        room_key_raw = str(room).strip()
        room_key = canonical_room_name(room_key_raw)
        rec["room_canonical"] = room_key
        rows.append(rec)
        totals = room_totals.setdefault(room_key, {"watts": 0.0, "kwh_per_year": 0.0})
        if isinstance(rec["watts"], (int, float)):
            totals["watts"] += float(rec["watts"])
        if isinstance(rec["kwh_per_year"], (int, float)):
            totals["kwh_per_year"] += float(rec["kwh_per_year"])

        r += 1

    return rows, room_totals


def build_room_dimension_summary(demand_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Aggregate usable room geometry/envelope metrics from Heating Demand rows."""

    by_room: dict[str, dict[str, Any]] = {}

    def _ensure(room: str) -> dict[str, Any]:
        return by_room.setdefault(
            room,
            {
                "room": room,
                "floor_area_m2": None,
                "ceiling_area_m2": None,
                "ceiling_height_m": None,
                "roof_area_m2": 0.0,
                "window_area_m2": 0.0,
                "velux_area_m2": 0.0,
                "external_wall_area_m2": 0.0,
                "internal_wall_area_m2": 0.0,
                "external_floor_area_m2": 0.0,
                "records_used": 0,
            },
        )

    for row in demand_rows:
        room = row.get("room_canonical") or row.get("room")
        if not isinstance(room, str):
            continue
        rec = _ensure(room)
        item = str(row.get("item") or "").strip().lower()
        area = row.get("area_m2")
        h = row.get("ceiling_height")

        if isinstance(h, (int, float)) and rec["ceiling_height_m"] is None:
            rec["ceiling_height_m"] = float(h)

        if isinstance(area, (int, float)):
            a = float(area)
            if item == "ceiling" and rec["ceiling_area_m2"] is None:
                rec["ceiling_area_m2"] = a
            if item in ("floor", "external floor") and rec["floor_area_m2"] is None:
                rec["floor_area_m2"] = a
            if item == "roof":
                rec["roof_area_m2"] += a
            if "window" in item:
                rec["window_area_m2"] += a
            if "velux" in item:
                rec["velux_area_m2"] += a
            if item == "external wall":
                rec["external_wall_area_m2"] += a
            if item == "wall":
                rec["internal_wall_area_m2"] += a
            if item == "external floor":
                rec["external_floor_area_m2"] += a

        rec["records_used"] += 1

    # Fill floor area from ceiling where missing (common for upper floors)
    for rec in by_room.values():
        if rec["floor_area_m2"] is None and isinstance(rec["ceiling_area_m2"], float):
            rec["floor_area_m2"] = rec["ceiling_area_m2"]

    return sorted(by_room.values(), key=lambda x: x["room"])


def parse_radiator_table(ws, header_row: int) -> list[dict[str, Any]]:
    cols = {
        "room": 1,
        "radiator_n": 2,
        "width_mm": 3,
        "height_mm": 4,
        "area_m2": 5,
        "type": 6,
        "t50_w": 7,
        "radiator_c": 8,
        "target_c": 9,
        "watts_at_target": 10,
        "notes": 11,
        "accuracy": 12,
    }

    rows: list[dict[str, Any]] = []
    r = header_row + 1
    while r <= ws.max_row:
        room = _norm(ws.cell(r, cols["room"]).value)
        if room in (None, ""):
            r += 1
            continue
        if room == "Total":
            break

        rec = {k: _norm(ws.cell(r, c).value) for k, c in cols.items()}
        rec["source"] = {"file": HEATING_XLSX.name, "sheet": ws.title, "row": r}
        rows.append(rec)
        r += 1

    return rows


def list_scan_sources() -> list[str]:
    files: list[str] = []
    if SUNLIGHT_OVERVIEW.exists():
        files.append(str(SUNLIGHT_OVERVIEW))
    if SUNLIGHT_SCAN_DIR.exists():
        for p in sorted(SUNLIGHT_SCAN_DIR.glob("SUNLIGHT LOFTS LIMITED part*.jpg")):
            files.append(str(p))
        for p in sorted(SUNLIGHT_SCAN_DIR.glob("BAC_*.jpg")):
            files.append(str(p))
    return files


def scan_dimension_points() -> list[dict[str, Any]]:
    """Manually transcribed dimensional points from IMG_0829 + SUNLIGHT part scans."""

    return [
        {
            "group": "conservatory",
            "label": "width_nominal",
            "value": 5900,
            "unit": "mm",
            "source": "BAC_3 (Hardwood Conservatory Survey Details)",
            "confidence": "high",
        },
        {
            "group": "conservatory",
            "label": "width_overall",
            "value": 5980,
            "unit": "mm",
            "source": "BAC_1 (Conservatory Base Survey Details)",
            "confidence": "high",
        },
        {
            "group": "conservatory",
            "label": "projection_overall",
            "value": 3500,
            "unit": "mm",
            "source": "BAC_3 (Hardwood Conservatory Survey Details)",
            "confidence": "high",
        },
        {
            "group": "conservatory",
            "label": "projection_internal",
            "value": 3360,
            "unit": "mm",
            "source": "BAC_3 (Hardwood Conservatory Survey Details)",
            "confidence": "high",
        },
        {
            "group": "conservatory",
            "label": "base_wall_height",
            "value": 600,
            "unit": "mm",
            "source": "BAC_1/BAC_3",
            "confidence": "high",
        },
        {
            "group": "conservatory",
            "label": "front_frame_height_dpc_to_top",
            "value": 2400,
            "unit": "mm",
            "source": "BAC_2/BAC_3",
            "confidence": "high",
        },
        {
            "group": "conservatory",
            "label": "roof_pitch",
            "value": 12.6,
            "unit": "deg",
            "source": "BAC_3",
            "confidence": "medium",
        },
        {
            "group": "loft_beam_layout",
            "label": "overall_width",
            "value": 5800,
            "unit": "mm",
            "source": "IMG_0829 / SUNLIGHT part 1_001",
            "confidence": "high",
        },
        {
            "group": "loft_beam_layout",
            "label": "overall_depth",
            "value": 3450,
            "unit": "mm",
            "source": "IMG_0829 / SUNLIGHT part 1_001",
            "confidence": "high",
        },
        {
            "group": "section",
            "label": "headroom_marked",
            "value": 2200,
            "unit": "mm",
            "source": "IMG_0829 / SUNLIGHT part 1",
            "confidence": "high",
        },
        {
            "group": "section",
            "label": "stair_total_rise",
            "value": 2825,
            "unit": "mm",
            "source": "IMG_0829 / SUNLIGHT part 1",
            "confidence": "high",
        },
        {
            "group": "window_schedule",
            "label": "bedroom_1_window",
            "value": "1200x1200",
            "unit": "mm",
            "source": "SUNLIGHT part 3",
            "confidence": "medium",
        },
        {
            "group": "window_schedule",
            "label": "bedroom_1_window_2",
            "value": "550x780",
            "unit": "mm",
            "source": "SUNLIGHT part 3",
            "confidence": "medium",
        },
        {
            "group": "window_schedule",
            "label": "bedroom_2_windows",
            "value": "2x 550x780",
            "unit": "mm",
            "source": "SUNLIGHT part 3",
            "confidence": "medium",
        },
        {
            "group": "window_schedule",
            "label": "bedroom_2_cabrio",
            "value": "940x2520",
            "unit": "mm",
            "source": "SUNLIGHT part 3",
            "confidence": "medium",
        },
        {
            "group": "window_schedule",
            "label": "shower_room_window",
            "value": "500x500",
            "unit": "mm",
            "source": "SUNLIGHT part 3",
            "confidence": "medium",
        },
        {
            "group": "window_schedule",
            "label": "staircase_window",
            "value": "1200x1200",
            "unit": "mm",
            "source": "SUNLIGHT part 3",
            "confidence": "medium",
        },
    ]


def parse_dim_token_to_area_m2(token: str) -> float | None:
    """Parse strings like '1200x1200' or '2x 550x780' to area in m²."""

    s = token.lower().replace(" ", "")
    multiplier = 1
    if s.startswith("2x"):
        multiplier = 2
        s = s[2:]
    parts = s.split("x")
    if len(parts) != 2:
        return None
    try:
        w = float(parts[0])
        h = float(parts[1])
    except ValueError:
        return None
    return multiplier * (w * h) / 1_000_000.0


def parse_loft_heating_rows() -> list[dict[str, Any]]:
    """Extract loft-specific dimensions from the separate loft workbook."""

    if not LOFT_HEATING_XLSX.exists():
        return []

    wb = load_workbook(LOFT_HEATING_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    room_map = {
        "B1": "Aldora",
        "B2": "Elvina",
        "Shower": "Shower",
        "Hall": "Hall",
    }

    rows: list[dict[str, Any]] = []
    for r in range(1, ws.max_row + 1):
        room = _norm(ws.cell(r, 1).value)
        item = _norm(ws.cell(r, 2).value)
        if not isinstance(room, str) or not isinstance(item, str):
            continue
        room_canonical = room_map.get(room.strip(), room.strip())
        rows.append(
            {
                "room": room,
                "room_canonical": room_canonical,
                "item": item,
                "meters": _norm(ws.cell(r, 3).value),
                "area_m2": _norm(ws.cell(r, 4).value),
                "u_value": _norm(ws.cell(r, 5).value),
                "watts": _norm(ws.cell(r, 6).value),
                "source": {
                    "file": str(LOFT_HEATING_XLSX),
                    "sheet": ws.title,
                    "row": r,
                },
            }
        )

    return rows


def bca_dimension_points() -> list[dict[str, Any]]:
    """Manually transcribed dimensional points from neighbour/BCA 2025 pack."""

    return [
        {
            "label": "min_clear_headroom",
            "value": 2000,
            "unit": "mm",
            "source": f"{BCA_EXISTING_PDF.name}: text note 'Maintain a clear 2.0m headroom'",
            "confidence": "high",
        },
        {
            "label": "first_floor_ffl",
            "value": 2800,
            "unit": "mm",
            "source": f"{BCA_EXISTING_PDF.name}: 'First FFL + 2.800'",
            "confidence": "high",
        },
        {
            "label": "second_floor_ffl",
            "value": 5520,
            "unit": "mm",
            "source": f"{BCA_EXISTING_PDF.name}: 'Second FFL + 5.520'",
            "confidence": "high",
        },
        {
            "label": "eaves_height",
            "value": 2800,
            "unit": "mm",
            "source": f"{BCA_SECTION_PDF.name}: 'eaves 2800'",
            "confidence": "high",
        },
        {
            "label": "velux_c02",
            "value": "550x780",
            "unit": "mm",
            "source": f"{BCA_EXISTING_PDF.name}: window type schedule 'GGL-C02 550 x 780'",
            "confidence": "medium",
        },
        {
            "label": "velux_s06",
            "value": "1180x1140",
            "unit": "mm",
            "source": f"{BCA_EXISTING_PDF.name}: window type schedule 'GGL-S06 1180 x 1140'",
            "confidence": "medium",
        },
    ]


def index_room_dims(room_dimension_summary: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {str(r["room"]): r for r in room_dimension_summary}


def build_loft_dimension_deltas(
    scan_dims: list[dict[str, Any]],
    bca_dims: list[dict[str, Any]],
    room_dimension_summary: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    by_scan = {d["label"]: d for d in scan_dims}
    by_bca = {d["label"]: d for d in bca_dims}
    rooms = index_room_dims(room_dimension_summary)

    rows: list[dict[str, Any]] = []

    def add(metric: str, sunlight: Any, bca: Any, xlsx: Any, unit: str, source_sun: str, source_bca: str, source_xlsx: str, confidence: str = "medium") -> None:
        d_sb = float(sunlight) - float(bca) if isinstance(sunlight, (int, float)) and isinstance(bca, (int, float)) else None
        d_sx = float(sunlight) - float(xlsx) if isinstance(sunlight, (int, float)) and isinstance(xlsx, (int, float)) else None
        rows.append(
            {
                "metric": metric,
                "sunlight_value": sunlight,
                "bca_value": bca,
                "xlsx_value": xlsx,
                "delta_sunlight_minus_bca": d_sb,
                "delta_sunlight_minus_xlsx": d_sx,
                "unit": unit,
                "confidence": confidence,
                "source_sunlight": source_sun,
                "source_bca": source_bca,
                "source_xlsx": source_xlsx,
            }
        )

    # Heights
    add(
        "loft_headroom",
        by_scan.get("headroom_marked", {}).get("value"),
        by_bca.get("min_clear_headroom", {}).get("value"),
        (rooms.get("Aldora", {}).get("ceiling_height_m") or 0) * 1000 if rooms.get("Aldora", {}).get("ceiling_height_m") is not None else None,
        "mm",
        by_scan.get("headroom_marked", {}).get("source", ""),
        by_bca.get("min_clear_headroom", {}).get("source", ""),
        "Heating Demand / Aldora ceiling height",
        "high",
    )

    bca_rise = None
    if isinstance(by_bca.get("first_floor_ffl", {}).get("value"), (int, float)) and isinstance(by_bca.get("second_floor_ffl", {}).get("value"), (int, float)):
        bca_rise = float(by_bca["second_floor_ffl"]["value"]) - float(by_bca["first_floor_ffl"]["value"])

    add(
        "loft_total_rise",
        by_scan.get("stair_total_rise", {}).get("value"),
        bca_rise,
        None,
        "mm",
        by_scan.get("stair_total_rise", {}).get("source", ""),
        f"{by_bca.get('first_floor_ffl', {}).get('source', '')} ; {by_bca.get('second_floor_ffl', {}).get('source', '')}",
        "n/a",
        "medium",
    )

    # Window areas
    sunlight_aldora = (
        (parse_dim_token_to_area_m2(str(by_scan.get("bedroom_1_window", {}).get("value", ""))) or 0)
        + (parse_dim_token_to_area_m2(str(by_scan.get("bedroom_1_window_2", {}).get("value", ""))) or 0)
    )
    bca_aldora = (
        (parse_dim_token_to_area_m2(str(by_bca.get("velux_s06", {}).get("value", ""))) or 0)
        + (parse_dim_token_to_area_m2(str(by_bca.get("velux_c02", {}).get("value", ""))) or 0)
    )
    add(
        "aldora_window_total_area",
        sunlight_aldora,
        bca_aldora,
        rooms.get("Aldora", {}).get("window_area_m2"),
        "m2",
        "SUNLIGHT part 3 bedroom_1_window + bedroom_1_window_2",
        "BCA window schedule types S06 + C02",
        "Heating Demand / Aldora windows",
        "medium",
    )

    sunlight_elvina_velux = parse_dim_token_to_area_m2(str(by_scan.get("bedroom_2_windows", {}).get("value", "")))
    bca_elvina_velux = 2 * (parse_dim_token_to_area_m2(str(by_bca.get("velux_c02", {}).get("value", ""))) or 0)
    add(
        "elvina_velux_pair_area",
        sunlight_elvina_velux,
        bca_elvina_velux,
        rooms.get("Elvina", {}).get("velux_area_m2"),
        "m2",
        "SUNLIGHT part 3 bedroom_2_windows",
        "BCA window schedule type C02 (x2)",
        "Heating Demand / Elvina velux",
        "medium",
    )

    sunlight_shower = (
        (parse_dim_token_to_area_m2(str(by_scan.get("shower_room_window", {}).get("value", ""))) or 0)
        + (parse_dim_token_to_area_m2("550x780") or 0)
    )
    add(
        "shower_window_total_area",
        sunlight_shower,
        parse_dim_token_to_area_m2(str(by_bca.get("velux_c02", {}).get("value", ""))),
        rooms.get("Shower", {}).get("window_area_m2"),
        "m2",
        "SUNLIGHT part 3 shower_room_window + inferred C02",
        "BCA window schedule type C02",
        "Heating Demand / Shower windows",
        "low",
    )

    return rows


def build_dimension_cross_reference(
    scan_dims: list[dict[str, Any]],
    room_dimension_summary: list[dict[str, Any]],
    conservatory_dims: dict[str, Any],
) -> list[dict[str, Any]]:
    by_label = {d["label"]: d for d in scan_dims}
    rooms = index_room_dims(room_dimension_summary)

    rows: list[dict[str, Any]] = []

    def add_row(
        metric: str,
        scan_value: Any,
        xlsx_value: Any,
        unit: str,
        comparison: str,
        source_scan: str,
        source_xlsx: str,
    ) -> None:
        delta = None
        if isinstance(scan_value, (int, float)) and isinstance(xlsx_value, (int, float)):
            delta = float(scan_value) - float(xlsx_value)
        rows.append(
            {
                "metric": metric,
                "scan_value": scan_value,
                "xlsx_value": xlsx_value,
                "delta": delta,
                "unit": unit,
                "comparison": comparison,
                "source_scan": source_scan,
                "source_xlsx": source_xlsx,
            }
        )

    # Conservatory footprint: BAC width x projection vs XLSX floor area.
    width = conservatory_dims.get("width_overall")
    proj = conservatory_dims.get("projection_overall")
    floor_area_scan = None
    if isinstance(width, (int, float)) and isinstance(proj, (int, float)):
        floor_area_scan = (float(width) * float(proj)) / 1_000_000.0
    conservatory_xlsx = rooms.get("Conservatory", {}).get("floor_area_m2")
    add_row(
        "conservatory_floor_area",
        floor_area_scan,
        conservatory_xlsx,
        "m2",
        "exact-ish",
        "BAC dimensions",
        "Heating Demand / Conservatory external floor",
    )

    # Derived dimensional cross-checks using XLSX area.
    if isinstance(conservatory_xlsx, (int, float)) and isinstance(width, (int, float)):
        proj_from_xlsx = (float(conservatory_xlsx) * 1_000_000.0) / float(width)
        add_row(
            "conservatory_projection_from_xlsx_area",
            proj,
            proj_from_xlsx,
            "mm",
            "derived",
            "BAC dimensions",
            "Derived from Heating Demand area and BAC width_overall",
        )

    if isinstance(conservatory_xlsx, (int, float)) and isinstance(proj, (int, float)):
        width_from_xlsx = (float(conservatory_xlsx) * 1_000_000.0) / float(proj)
        add_row(
            "conservatory_width_from_xlsx_area",
            width,
            width_from_xlsx,
            "mm",
            "derived",
            "BAC dimensions",
            "Derived from Heating Demand area and BAC projection_overall",
        )

    # Loft dimensions cross-reference.
    scan_headroom = by_label.get("headroom_marked", {}).get("value")
    loft_ceiling_h = rooms.get("Aldora", {}).get("ceiling_height_m")
    loft_ceiling_h_mm = None
    if isinstance(loft_ceiling_h, (int, float)):
        loft_ceiling_h_mm = float(loft_ceiling_h) * 1000.0
    add_row(
        "loft_headroom_vs_ceiling_height",
        scan_headroom,
        loft_ceiling_h_mm,
        "mm",
        "proxy",
        "IMG_0829 section",
        "Heating Demand / Aldora ceiling height",
    )

    # Window schedule comparisons.
    b1_main = parse_dim_token_to_area_m2(str(by_label.get("bedroom_1_window", {}).get("value", "")))
    b1_aux = parse_dim_token_to_area_m2(str(by_label.get("bedroom_1_window_2", {}).get("value", "")))
    b1_total = None
    if isinstance(b1_main, float) and isinstance(b1_aux, float):
        b1_total = b1_main + b1_aux

    add_row(
        "aldora_window_area_total",
        b1_total,
        rooms.get("Aldora", {}).get("window_area_m2"),
        "m2",
        "exact-ish",
        "SUNLIGHT part 3 window schedule",
        "Heating Demand / Aldora windows",
    )

    b2_cabrio = parse_dim_token_to_area_m2(str(by_label.get("bedroom_2_cabrio", {}).get("value", "")))
    b2_velux_pair = parse_dim_token_to_area_m2(str(by_label.get("bedroom_2_windows", {}).get("value", "")))

    add_row(
        "elvina_window_area_cabrio",
        b2_cabrio,
        rooms.get("Elvina", {}).get("window_area_m2"),
        "m2",
        "exact-ish",
        "SUNLIGHT part 3 window schedule",
        "Heating Demand / Elvina windows",
    )

    add_row(
        "elvina_velux_area_pair",
        b2_velux_pair,
        rooms.get("Elvina", {}).get("velux_area_m2"),
        "m2",
        "exact-ish",
        "SUNLIGHT part 3 window schedule",
        "Heating Demand / Elvina velux",
    )

    shower_main = parse_dim_token_to_area_m2(str(by_label.get("shower_room_window", {}).get("value", "")))
    shower_aux = parse_dim_token_to_area_m2("550x780")
    shower_total = None
    if isinstance(shower_main, float) and isinstance(shower_aux, float):
        shower_total = shower_main + shower_aux

    add_row(
        "shower_window_area_total",
        shower_total,
        rooms.get("Shower", {}).get("window_area_m2"),
        "m2",
        "exact-ish",
        "SUNLIGHT part 3 window schedule",
        "Heating Demand / Shower windows",
    )

    # User-confirmed overrides where spreadsheet reflects later changes better than drawings.
    xlsx_preferred_metrics = {
        "aldora_window_area_total": "Window values changed after drawing issue; use XLSX as canonical.",
        "shower_window_area_total": "Window values changed after drawing issue; use XLSX as canonical.",
    }

    # Evaluate simple status
    for row in rows:
        metric = row.get("metric")
        if metric in xlsx_preferred_metrics:
            row["status"] = "xlsx_override"
            row["canonical_source"] = "xlsx"
            row["canonical_value"] = row.get("xlsx_value")
            row["override_reason"] = xlsx_preferred_metrics[metric]
            continue

        if isinstance(row["delta"], float):
            ad = abs(row["delta"])
            if row["unit"] == "mm":
                row["status"] = "match" if ad <= 50 else "check"
            else:
                row["status"] = "match" if ad <= 0.10 else "check"
            row["canonical_source"] = "scan"
            row["canonical_value"] = row.get("scan_value")
            row["override_reason"] = ""
        else:
            row["status"] = "insufficient_data"
            row["canonical_source"] = "unknown"
            row["canonical_value"] = ""
            row["override_reason"] = ""

    return rows


def canonical_dimension_rows_from_geometry() -> list[dict[str, Any]]:
    if not THERMAL_GEOMETRY_JSON.exists():
        return []
    geo = json.loads(THERMAL_GEOMETRY_JSON.read_text())
    rows: list[dict[str, Any]] = []

    for r in geo.get("rooms", []):
        room = r["name"]
        rows.append({"entity": room, "kind": "room", "metric": "floor_area", "value": r["floor_area"], "unit": "m2", "source": str(THERMAL_GEOMETRY_JSON)})
        rows.append({"entity": room, "kind": "room", "metric": "ceiling_height", "value": r["ceiling_height"], "unit": "m", "source": str(THERMAL_GEOMETRY_JSON)})

        for e in r.get("external_fabric", []):
            rows.append({
                "entity": room,
                "kind": "external_fabric",
                "metric": e["description"],
                "value": e["area"],
                "unit": "m2",
                "source": str(THERMAL_GEOMETRY_JSON),
            })

    for c in geo.get("connections", []):
        rows.append({
            "entity": f"{c['room_a']}->{c['room_b']}",
            "kind": "connection",
            "metric": "ua",
            "value": c["ua"],
            "unit": "W/K",
            "source": str(THERMAL_GEOMETRY_JSON),
        })

    for d in geo.get("doorways", []):
        rows.append({
            "entity": f"{d['room_a']}->{d['room_b']}",
            "kind": "doorway",
            "metric": "width",
            "value": d["width"],
            "unit": "m",
            "source": str(THERMAL_GEOMETRY_JSON),
        })
        rows.append({
            "entity": f"{d['room_a']}->{d['room_b']}",
            "kind": "doorway",
            "metric": "height",
            "value": d["height"],
            "unit": "m",
            "source": str(THERMAL_GEOMETRY_JSON),
        })

    return rows


def compute_ewi_area_quantification(geo: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Compute net EWI area using agreed rule: gross section envelope - windows."""

    target_rooms = {"front", "hall", "jackcarol", "office"}
    gross_wall = 0.0
    subtract_windows = 0.0
    breakdown: list[dict[str, Any]] = []

    for room in geo.get("rooms", []):
        name = str(room.get("name", "")).lower()
        if name not in target_rooms:
            continue
        for e in room.get("external_fabric", []):
            desc = str(e.get("description", ""))
            area = float(e.get("area", 0.0))
            is_wall = "external wall" in desc.lower()
            is_window = "window" in desc.lower() and "loft windows" not in desc.lower()
            if is_wall:
                gross_wall += area
            if is_window:
                subtract_windows += area
            breakdown.append(
                {
                    "room": name,
                    "element": desc,
                    "area_m2": area,
                    "counts_in_gross_wall": is_wall,
                    "counts_as_window_subtraction": is_window,
                }
            )

    net = gross_wall - subtract_windows
    summary = {
        "target_rooms": sorted(target_rooms),
        "gross_section_envelope_m2": round(gross_wall, 4),
        "window_subtraction_m2": round(subtract_windows, 4),
        "net_ewi_area_m2": round(net, 4),
        "formula": "net_ewi_area_m2 = gross_section_envelope_m2 - window_subtraction_m2",
        "status": "locked",
        "source": "Computed from data/canonical/thermal_geometry.json external_fabric records",
    }
    return summary, breakdown


def _room_item_candidates_from_desc(desc: str) -> list[str]:
    dl = desc.lower()
    if "external wall" in dl:
        return ["external wall", "wall"]
    if "roof" in dl:
        return ["roof"]
    if "velux" in dl:
        return ["velux"]
    if "window" in dl:
        return ["window", "windows"]
    if "ground floor" in dl or "floor" in dl:
        return ["external floor", "floor", "ceiling"]
    return []


def _find_xlsx_source_row(
    xlsx_rows: list[dict[str, Any]],
    room_name: str,
    item_candidates: list[str],
    value: float,
    value_field: str = "area_m2",
    tolerance: float = 1e-6,
) -> dict[str, Any] | None:
    room_norm = room_name.strip().lower()
    cands = [c.lower() for c in item_candidates]
    for row in xlsx_rows:
        rr = str(row.get("room_canonical") or "").strip().lower()
        it = str(row.get("item") or "").strip().lower()
        if rr != room_norm:
            continue
        if cands and it not in cands:
            continue
        v = row.get(value_field)
        if isinstance(v, (int, float)) and abs(float(v) - float(value)) <= tolerance:
            return row
    return None


def _xlsx_ref(row: dict[str, Any], col: str) -> str:
    src = row.get("source", {}) if isinstance(row, dict) else {}
    file = src.get("file", "")
    sheet = src.get("sheet", "")
    r = src.get("row")
    return f"{file}::{sheet}!{col}{r}" if r is not None else f"{file}::{sheet}"


def _build_geometry_provenance(
    geo: dict[str, Any],
    demand_rows: list[dict[str, Any]],
    loft_rows: list[dict[str, Any]],
    current_rads: list[dict[str, Any]],
) -> dict[str, Any]:
    """Build field-level provenance for every canonical geometry leaf path."""

    all_rows = [*demand_rows, *loft_rows]
    prov: dict[str, Any] = {}

    rad_by_room: dict[str, list[dict[str, Any]]] = {}
    for r in current_rads:
        room = canonical_room_name(str(r.get("room") or "")).strip().lower()
        rad_by_room.setdefault(room, []).append(r)

    for ri, room in enumerate(geo.get("rooms", [])):
        room_name = str(room.get("name", ""))
        room_title = room_name.replace("jackcarol", "Jack & Carol").title()

        # Room-level fields
        floor_row = _find_xlsx_source_row(all_rows, room_title, ["floor", "external floor", "ceiling"], float(room.get("floor_area", 0)), "area_m2", 1e-3)
        prov[f"rooms[{ri}].floor_area"] = {
            "source_type": "xlsx" if floor_row else "canonical_manual",
            "source_ref": _xlsx_ref(floor_row, "E") if floor_row else "manual/canonical_geometry",
        }

        ch_row = _find_xlsx_source_row(all_rows, room_title, [], float(room.get("ceiling_height", 0)), "ceiling_height", 1e-3)
        prov[f"rooms[{ri}].ceiling_height"] = {
            "source_type": "xlsx" if ch_row else "canonical_manual",
            "source_ref": _xlsx_ref(ch_row, "D") if ch_row else "manual/canonical_geometry",
        }

        prov[f"rooms[{ri}].name"] = {"source_type": "manual_mapping", "source_ref": "house_inventory.room_alias_map"}
        prov[f"rooms[{ri}].floor"] = {"source_type": "manual_mapping", "source_ref": "house_inventory.room_alias_map"}
        prov[f"rooms[{ri}].construction"] = {"source_type": "model_assumption", "source_ref": "model/house.py construction classes"}
        prov[f"rooms[{ri}].sensor"] = {"source_type": "sensor_map", "source_ref": "model/house.py build_sensor_map"}
        prov[f"rooms[{ri}].ventilation_ach"] = {"source_type": "calibrated", "source_ref": "Night 1/2 fit (model/house.py)"}
        prov[f"rooms[{ri}].heat_recovery"] = {"source_type": "equipment_spec", "source_ref": "Bathroom MVHR spec"}
        prov[f"rooms[{ri}].overnight_occupants"] = {"source_type": "occupancy_assumption", "source_ref": "model/house.py constants"}

        # Radiators
        room_rads = rad_by_room.get(room_title.lower(), [])
        for rj, rad in enumerate(room.get("radiators", [])):
            src = room_rads[rj] if rj < len(room_rads) else None
            prov[f"rooms[{ri}].radiators[{rj}].t50"] = {
                "source_type": "xlsx" if src else "canonical_manual",
                "source_ref": _xlsx_ref(src, "G") if src else "manual/canonical_geometry",
            }
            prov[f"rooms[{ri}].radiators[{rj}].pipe"] = {
                "source_type": "hydraulic_topology",
                "source_ref": "AGENTS.md pipe topology",
            }
            prov[f"rooms[{ri}].radiators[{rj}].active"] = {
                "source_type": "observed_state",
                "source_ref": "AGENTS.md radiator status",
            }

        # External fabric
        for ej, elem in enumerate(room.get("external_fabric", [])):
            desc = str(elem.get("description", ""))
            area = float(elem.get("area", 0))
            row = _find_xlsx_source_row(all_rows, room_title, _room_item_candidates_from_desc(desc), area, "area_m2", 1e-3)
            prov[f"rooms[{ri}].external_fabric[{ej}].description"] = {
                "source_type": "manual_label",
                "source_ref": "canonical_geometry description",
            }
            prov[f"rooms[{ri}].external_fabric[{ej}].area"] = {
                "source_type": "xlsx" if row else "scan_or_manual",
                "source_ref": _xlsx_ref(row, "E" if str(row.get("source", {}).get("file", "")).endswith("Heating needs for the house.xlsx") else "D") if row else "scan_dimension_points / canonical_geometry",
            }
            u_row = row
            prov[f"rooms[{ri}].external_fabric[{ej}].u_value"] = {
                "source_type": "xlsx" if u_row and isinstance(u_row.get("u_value"), (int, float)) else "u_value_library",
                "source_ref": _xlsx_ref(u_row, "F" if str(u_row.get("source", {}).get("file", "")).endswith("Heating needs for the house.xlsx") else "E") if u_row else "config/spec assumptions",
            }
            prov[f"rooms[{ri}].external_fabric[{ej}].to_ground"] = {
                "source_type": "manual_classification",
                "source_ref": "canonical_geometry element classification",
            }

        # Solar
        for sj, _solar in enumerate(room.get("solar", [])):
            for field_name in ("area", "orientation", "tilt", "g_value", "shading"):
                prov[f"rooms[{ri}].solar[{sj}].{field_name}"] = {
                    "source_type": "solar_calibration",
                    "source_ref": "model/house.py solar gain calibration (26 Mar 2026)",
                }

    for ci, conn in enumerate(geo.get("connections", [])):
        desc = str(conn.get("description", ""))
        conn_source_type = "model_definition"
        conn_source_ref = "canonical geometry migration from model/house.py"
        if desc == "Internal wall":
            conn_source_type = "xlsx_internal_wall_constrained"
            conn_source_ref = "Adjusted to match room_dimension_summary.internal_wall_area_m2 totals (plan-derived spreadsheet)"
        for field_name in ("room_a", "room_b", "ua", "description"):
            prov[f"connections[{ci}].{field_name}"] = {
                "source_type": conn_source_type,
                "source_ref": conn_source_ref,
            }

    for di, _door in enumerate(geo.get("doorways", [])):
        for field_name in ("room_a", "room_b", "width", "height", "state"):
            prov[f"doorways[{di}].{field_name}"] = {
                "source_type": "model_definition",
                "source_ref": "canonical geometry migration from model/house.py",
            }

    return prov


def write_csv(path: Path, rows: list[dict[str, Any]], field_order: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=field_order)
        w.writeheader()
        for row in rows:
            out = dict(row)
            src = out.pop("source", None)
            if isinstance(src, dict):
                out["source_file"] = src.get("file")
                out["source_sheet"] = src.get("sheet")
                out["source_row"] = src.get("row")
            elif src is not None:
                out["source"] = src
            w.writerow(out)


def main() -> None:
    wb = load_workbook(HEATING_XLSX, data_only=True)

    ws_demand = wb["Heating Demand"]
    ws_rads = wb["Current Radiators"]

    demand_rows, demand_totals = parse_heating_demand(ws_demand)
    room_dimension_summary = build_room_dimension_summary(demand_rows)
    scan_sources = list_scan_sources()
    scan_dims = scan_dimension_points()
    bca_dims = bca_dimension_points()
    loft_rows = parse_loft_heating_rows()
    canonical_dimension_rows = canonical_dimension_rows_from_geometry()
    geo = json.loads(THERMAL_GEOMETRY_JSON.read_text()) if THERMAL_GEOMETRY_JSON.exists() else {}
    ewi_summary, ewi_breakdown = compute_ewi_area_quantification(geo)
    loft_dimension_deltas = build_loft_dimension_deltas(scan_dims, bca_dims, room_dimension_summary)
    conservatory_dims = {
        "width_overall": 5980,
        "width_nominal": 5900,
        "projection_overall": 3500,
        "projection_internal": 3360,
        "base_wall_height": 600,
        "front_frame_height_dpc_to_top": 2400,
    }
    dimension_cross_ref = build_dimension_cross_reference(
        scan_dims,
        room_dimension_summary,
        conservatory_dims,
    )

    rad_headers = find_headers(ws_rads, "Room", "Radiator")
    if not rad_headers:
        raise RuntimeError("Could not find radiator table headers.")

    radiator_tables: list[dict[str, Any]] = []
    for i, header_row in enumerate(rad_headers, start=1):
        rows = parse_radiator_table(ws_rads, header_row)
        label = f"table_{i}"
        if i == 1:
            label = "current"
        elif i == 2:
            label = "change"
        elif i == 3:
            label = "extras"
        radiator_tables.append(
            {
                "label": label,
                "header_row": header_row,
                "records": rows,
            }
        )

    current_rads = next((t["records"] for t in radiator_tables if t["label"] == "current"), [])
    change_rads = next((t["records"] for t in radiator_tables if t["label"] == "change"), [])

    # Inject field-level provenance into canonical runtime geometry.
    if geo:
        geo["provenance"] = _build_geometry_provenance(geo, demand_rows, loft_rows, current_rads)
        THERMAL_GEOMETRY_JSON.write_text(json.dumps(geo, indent=2) + "\n", encoding="utf-8")
        canonical_dimension_rows = canonical_dimension_rows_from_geometry()

    # Source hierarchy and manual corrections from user review of scans (2026-03-27).
    inventory = {
        "schema_version": "1.0.0",
        "generated_at": "2026-03-27",
        "source_hierarchy": [
            {
                "domain": "conservatory",
                "priority": [
                    "BAC survey/order scans (1999) - definitive",
                    "Sunlight Lofts drawing (2009) - contextual",
                ],
            },
            {
                "domain": "loft",
                "priority": [
                    "Sunlight overview IMG_0829 + 'SUNLIGHT LOFTS LIMITED part*' scans - primary dimensional source",
                    "Cross-reference dimensions against XLSX data",
                    "Neighbour/BCA pack (2025) - secondary cross-check only",
                ],
            },
            {
                "domain": "thermal/radiator calculations",
                "priority": [
                    "Heating needs for the house.xlsx",
                    "Utility - Gas Electric-Jack_Laptop.xlsx (supporting energy history)",
                ],
            },
        ],
        "dimensional_data_policy": {
            "primary_visual_sources": [
                "IMG_0829.jpeg (full-sheet Sunlight overview)",
                "/mnt/c/Users/jackc/AppData/Roaming/PFU/ScanSnap Home/ScanSnap Home/SUNLIGHT LOFTS LIMITED part*",
                "/mnt/c/Users/jackc/AppData/Roaming/PFU/ScanSnap Home/ScanSnap Home/BAC_*",
            ],
            "cross_reference_sources": [
                "Heating needs for the house.xlsx",
                "Utility - Gas Electric-Jack_Laptop.xlsx",
            ],
            "rule": "Use Sunlight overview + part scans for loft geometry, BAC_* scans for conservatory geometry, and cross-reference dimensions with XLSX values.",
            "source": "User instruction, 2026-03-27",
        },
        "scan_sources_used": scan_sources,
        "scan_dimension_points": scan_dims,
        "bca_dimension_points": bca_dims,
        "confidence_scale": {
            "high": "Confirmed by explicit user instruction and/or definitive source docs",
            "medium": "Strongly supported but may rely on interpretation or partial scan legibility",
            "low": "Inferred or pending confirmation",
        },
        "room_alias_map": {
            "sunlight_ground_floor": {
                "lounge": "leather",
                "living_room": "front",
                "kitchen": "kitchen",
                "hall": "hall",
                "conservatory": "conservatory",
            },
            "sunlight_first_floor": {
                "smallest_bedroom": "office",
                "largest_bedroom_with_bay": "jackcarol",
                "other_bedroom": "sterling",
                "bath_room": "bathroom",
                "wc": "wc",
            },
            "sunlight_loft_floor": {
                "bedroom_1": "aldora",
                "bedroom_2": "elvina",
                "shower_room": "shower",
                "landing_or_hall": "landing",
            },
        },
        "user_corrections": [
            {
                "id": "loft_bedroom2_door_orientation",
                "statement": "Bedroom 2 (Elvina) door is at position 1, opening down on plan (not opening right).",
                "source": "User instruction, 2026-03-27",
            },
            {
                "id": "loft_bedroom2_pencil3_included",
                "statement": "Pencilled area '3' is included within Bedroom 2 (Elvina), not excluded.",
                "source": "User instruction, 2026-03-27",
            },
            {
                "id": "aldora_window_override",
                "statement": "Bedroom 1 (Aldora) window follows Excel spreadsheet spec where drawing differs.",
                "source": "User instruction, 2026-03-27",
            },
            {
                "id": "window_area_xlsx_preferred_for_changed_items",
                "statement": "For Aldora and Shower window totals, spreadsheet values are more accurate than the drawing because those items were changed.",
                "source": "User instruction, 2026-03-27",
            },
            {
                "id": "radiator_change_intent",
                "statement": "Aldora 'Change' table entry is a possible future swap from current towel radiator to DP DF; not yet implemented.",
                "source": "User instruction, 2026-03-27",
            },
            {
                "id": "radiator_extra_not_implemented",
                "statement": "Radiators listed in 'Extra' (Landing vertical + Kitchen kickspace) have not been implemented and will not be implemented.",
                "source": "User instruction, 2026-03-27",
            },
        ],
        "radiator_implementation_status": {
            "current_table": "implemented",
            "change_table": "planned_candidate_only_not_implemented",
            "extras_table": "not_used",
            "dimension_labels": "width/height values are correct as-entered in current and change tables",
        },
        "radiator_table_policy": {
            "current": {
                "status": "canonical",
                "width_height": "correct",
                "usage": "used",
                "basis": "User confirmation, 2026-03-27",
            },
            "change": {
                "status": "possible_future_change",
                "width_height": "correct",
                "usage": "scenario_only",
                "basis": "User confirmation, 2026-03-27",
            },
            "extras": {
                "status": "not_used",
                "width_height": "as listed",
                "usage": "ignore_in_programs",
                "basis": "User confirmation, 2026-03-27",
            },
        },
        "dimensional_consistency": {
            "sunlight_vs_neighbour": "expected_close_match",
            "sunlight_vs_excel": "expected_close_match",
            "basis": "User instruction, 2026-03-27",
        },
        "conservatory_definitive": {
            "source": "BAC survey/order scans (1999)",
            "dimensions_mm": conservatory_dims,
            "roof": {"pitch_deg": 12.6},
            "notes": [
                "Finished floor level set ~150mm below existing kitchen door cill (survey note).",
                "Drainage/basework notes include new gully/soakaway references.",
            ],
        },
        "loft_dimensions_from_sunlight_mm": {
            "beam_layout_width": 5800,
            "beam_layout_depth": 3450,
            "section_headroom": 2200,
            "section_total_rise": 2825,
            "source": "Sunlight full sheet photo IMG_0829 + part scans",
        },
        "ewi_extent_guidance": {
            "max_extent_reference": "IMG_0829 section view (ground + first floors)",
            "interpretation": "Section envelope at ground and first floors is the maximum potential EWI replacement extent.",
            "net_area_adjustment": "Ground/first-floor window areas must be subtracted from gross section extent.",
            "status": "locked",
            "source": "User instruction, 2026-03-27",
            "quantification": ewi_summary,
            "breakdown_records": len(ewi_breakdown),
        },
        "xlsx_cross_reference": {
            "heating_demand": {
                "file": HEATING_XLSX.name,
                "sheet": "Heating Demand",
                "header_row": find_headers(ws_demand, "Room", "Item")[0],
                "records": len(demand_rows),
            },
            "radiator_tables": [
                {
                    "label": t["label"],
                    "file": HEATING_XLSX.name,
                    "sheet": "Current Radiators",
                    "header_row": t["header_row"],
                    "records": len(t["records"]),
                }
                for t in radiator_tables
            ],
            "loft_heating": {
                "file": str(LOFT_HEATING_XLSX),
                "sheet": "Sheet1",
                "records": len(loft_rows),
                "status": "loaded" if loft_rows else "missing",
            },
        },
        "room_totals_from_heating_demand": demand_totals,
        "room_dimension_summary": room_dimension_summary,
        "dimension_cross_reference": dimension_cross_ref,
        "loft_dimension_deltas": loft_dimension_deltas,
        "ewi_area_breakdown": ewi_breakdown,
        "confidence_by_section": {
            "conservatory_definitive": {
                "level": "high",
                "basis": "User designated BAC docs as definitive + dimensions extracted from BAC scans",
            },
            "room_alias_map": {
                "level": "high",
                "basis": "Explicit room mapping provided by user for Sunlight first/ground/loft plans",
            },
            "user_corrections": {
                "level": "high",
                "basis": "Direct user corrections (door orientation, included area, Aldora window override)",
            },
            "loft_dimensions_from_sunlight_mm": {
                "level": "medium",
                "basis": "Read from scanned drawings/photo (IMG_0829 + parts), legibility may limit exactness",
            },
            "xlsx_cross_reference": {
                "level": "high",
                "basis": "Programmatic extraction from workbook with source sheet/row provenance",
            },
            "radiator_implementation_status": {
                "level": "high",
                "basis": "Direct user instruction on implemented vs not implemented tables",
            },
            "radiator_table_policy": {
                "level": "high",
                "basis": "Direct user confirmation of table usage and width/height correctness",
            },
            "room_totals_from_heating_demand": {
                "level": "high",
                "basis": "Direct aggregation from Heating Demand sheet rows",
            },
            "scan_dimension_points": {
                "level": "medium",
                "basis": "Manually transcribed from scan visuals; key dimensions high-confidence, some schedule entries medium-confidence pending OCR/cell pinning",
            },
            "dimension_cross_reference": {
                "level": "medium",
                "basis": "Computed comparison between scan-derived points and XLSX-derived room dimensions; some rows are proxy comparisons",
            },
            "dimensional_data_policy": {
                "level": "high",
                "basis": "Direct user instruction naming IMG_0829 + SUNLIGHT LOFTS LIMITED part* as dimensional source",
            },
            "dimensional_consistency": {
                "level": "medium",
                "basis": "Expectation from user statement; not yet independently dimension-matched end-to-end",
            },
            "ewi_extent_guidance": {
                "level": "high",
                "basis": "Direct user instruction tied to IMG_0829 section view",
            },
        },
        "open_questions": [
            {
                "id": "loft_dimension_full_validation",
                "status": "resolved",
                "confidence": "medium",
                "detail": "Loft delta table generated in loft_dimension_deltas.csv and embedded under loft_dimension_deltas.",
            },
            {
                "id": "ewi_net_area_quantification",
                "status": "resolved",
                "confidence": "high",
                "detail": "Net EWI area computed and locked in ewi_extent_guidance.quantification and ewi_area_breakdown.csv.",
            },
        ],
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # JSON inventory
    with (OUT_DIR / "house_inventory.json").open("w", encoding="utf-8") as f:
        json.dump(inventory, f, indent=2)
        f.write("\n")

    # CSV exports for programs
    write_csv(
        OUT_DIR / "heating_demand_items.csv",
        demand_rows,
        [
            "room",
            "room_canonical",
            "item",
            "meters",
            "ceiling_height",
            "area_m2",
            "u_value",
            "temp_diff_c",
            "watts",
            "kwh_per_year",
            "source_file",
            "source_sheet",
            "source_row",
        ],
    )

    write_csv(
        OUT_DIR / "room_dimension_summary.csv",
        room_dimension_summary,
        [
            "room",
            "floor_area_m2",
            "ceiling_area_m2",
            "ceiling_height_m",
            "roof_area_m2",
            "window_area_m2",
            "velux_area_m2",
            "external_wall_area_m2",
            "internal_wall_area_m2",
            "external_floor_area_m2",
            "records_used",
        ],
    )

    write_csv(
        OUT_DIR / "scan_dimension_points.csv",
        scan_dims,
        ["group", "label", "value", "unit", "source", "confidence"],
    )

    write_csv(
        OUT_DIR / "dimension_cross_reference.csv",
        dimension_cross_ref,
        [
            "metric",
            "scan_value",
            "xlsx_value",
            "delta",
            "unit",
            "comparison",
            "status",
            "canonical_source",
            "canonical_value",
            "override_reason",
            "source_scan",
            "source_xlsx",
        ],
    )

    write_csv(
        OUT_DIR / "canonical_dimensions.csv",
        canonical_dimension_rows,
        ["entity", "kind", "metric", "value", "unit", "source"],
    )

    write_csv(
        OUT_DIR / "loft_dimension_deltas.csv",
        loft_dimension_deltas,
        [
            "metric",
            "sunlight_value",
            "bca_value",
            "xlsx_value",
            "delta_sunlight_minus_bca",
            "delta_sunlight_minus_xlsx",
            "unit",
            "confidence",
            "source_sunlight",
            "source_bca",
            "source_xlsx",
        ],
    )

    write_csv(
        OUT_DIR / "ewi_area_breakdown.csv",
        ewi_breakdown,
        [
            "room",
            "element",
            "area_m2",
            "counts_in_gross_wall",
            "counts_as_window_subtraction",
        ],
    )

    provenance_rows = [
        {
            "path": k,
            "source_type": v.get("source_type", ""),
            "source_ref": v.get("source_ref", ""),
        }
        for k, v in sorted((geo.get("provenance") or {}).items())
    ]
    write_csv(
        OUT_DIR / "canonical_geometry_provenance.csv",
        provenance_rows,
        ["path", "source_type", "source_ref"],
    )

    rad_fields = [
        "room",
        "radiator_n",
        "width_mm",
        "height_mm",
        "area_m2",
        "type",
        "t50_w",
        "radiator_c",
        "target_c",
        "watts_at_target",
        "notes",
        "accuracy",
        "source_file",
        "source_sheet",
        "source_row",
    ]

    # Stable names for first two tables
    write_csv(OUT_DIR / "radiator_inventory_current.csv", current_rads, rad_fields)
    write_csv(OUT_DIR / "radiator_inventory_change.csv", change_rads, rad_fields)

    # Export every discovered radiator table for downstream programs.
    for t in radiator_tables:
        write_csv(
            OUT_DIR / f"radiator_inventory_{t['label']}.csv",
            t["records"],
            rad_fields,
        )

    print(f"Wrote inventory artifacts to {OUT_DIR}")
    print(f"  heating_demand_items.csv: {len(demand_rows)} rows")
    print(f"  room_dimension_summary.csv: {len(room_dimension_summary)} rows")
    print(f"  scan_dimension_points.csv: {len(scan_dims)} rows")
    print(f"  dimension_cross_reference.csv: {len(dimension_cross_ref)} rows")
    print(f"  canonical_dimensions.csv: {len(canonical_dimension_rows)} rows")
    print(f"  loft_dimension_deltas.csv: {len(loft_dimension_deltas)} rows")
    print(f"  ewi_area_breakdown.csv: {len(ewi_breakdown)} rows")
    print(f"  canonical_geometry_provenance.csv: {len(provenance_rows)} rows")
    for t in radiator_tables:
        print(f"  radiator_inventory_{t['label']}.csv: {len(t['records'])} rows")


if __name__ == "__main__":
    main()
